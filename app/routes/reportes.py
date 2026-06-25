"""
reportes.py
Blueprint de Reportes — accesible solo para roles 'administrador' y 'transcriptor'.

R1 – Alumnos por Curso
R2 – Alumnos por Curso con Calificaciones
R3 – Alumnos por Curso con Estado de Pagos
R4 – Estado de Cuenta de un Alumno
R5 – Top 3 mejores promedios por Curso
R6 – Resumen Financiero por Curso          (adicional)
R7 – Promedios y Tasa de Aprobación por Materia  (adicional)
R8 – Alumnos con Deuda o Abandono          (adicional)

Exportación: HTML en pantalla, Excel (.xlsx) y PDF (.pdf).
"""
import io
from datetime import datetime

from flask import (Blueprint, abort, render_template,
                   request, send_file)
from flask_login import current_user, login_required
from sqlalchemy import func

from app import db
from app.models import (Alumno, Asignado, Curso, Inscrito,
                        Materia, Nota, Pago, Profesor)

reportes_bp = Blueprint('reportes', __name__, url_prefix='/reportes')

_PRIM     = (124, 31, 62)   # #7C1F3E
_XL_HDR   = 'FF7C1F3E'
_XL_ALT   = 'FFFAE4EB'
_XL_GREEN = 'FFD4EDDA'
_XL_RED   = 'FFF8D7DA'
_XL_AMBER = 'FFFFF3CD'


# ── Guard ─────────────────────────────────────────────────────────────────────

def _guard():
    if not (current_user.has_role('administrador') or
            current_user.has_role('transcriptor')):
        abort(403)


# ── Helpers datos ─────────────────────────────────────────────────────────────

def _cursos():
    return Curso.query.order_by(Curso.gestion.desc(), Curso.paralelo).all()


def _alumnos():
    return (Alumno.query.filter_by(activo=True)
            .order_by(Alumno.paterno, Alumno.nombre).all())


def _nc(curso):
    if not curso:
        return '—'
    g = curso.grado.grado if curso.grado else ''
    return f"{g} – {curso.paralelo} ({curso.gestion})"


def _ins_q(cur_id):
    return (Inscrito.query
            .join(Alumno, Inscrito.alu_id == Alumno.id)
            .filter(Inscrito.cur_id == cur_id,
                    db.or_(Inscrito.inscrito == True,
                           Inscrito.reserva  == True)))


# ── Helpers Excel ─────────────────────────────────────────────────────────────

def _xs():
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    t = Side(style='thin', color='FFB0B0B0')
    b = Border(left=t, right=t, top=t, bottom=t)
    return {
        'hf': PatternFill('solid', fgColor=_XL_HDR),
        'af': PatternFill('solid', fgColor=_XL_ALT),
        'gf': PatternFill('solid', fgColor=_XL_GREEN),
        'rf': PatternFill('solid', fgColor=_XL_RED),
        'am': PatternFill('solid', fgColor=_XL_AMBER),
        'hft': Font(bold=True, color='FFFFFFFF', size=10),
        'bf':  Font(size=9),
        'bld': Font(bold=True, size=9),
        'ctr': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'lft': Alignment(horizontal='left',   vertical='center', wrap_text=True),
        'brd': b,
    }


def _xh(ws, cols, s):
    ws.append(cols)
    for c in ws[1]:
        c.fill = s['hf']; c.font = s['hft']
        c.border = s['brd']; c.alignment = s['ctr']


def _xrow(ws, s, fill=None):
    for c in ws[ws.max_row]:
        if fill: c.fill = fill
        c.font = s['bf']; c.border = s['brd']; c.alignment = s['lft']


def _xz(ws, s, start=2):
    for i, row in enumerate(ws.iter_rows(min_row=start)):
        for c in row:
            if i % 2 == 0: c.fill = s['af']
            c.font = s['bf']; c.border = s['brd']; c.alignment = s['lft']


def _xw(ws, lo=10, hi=42):
    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col), default=lo)
        ws.column_dimensions[col[0].column_letter].width = min(max(w+2, lo), hi)


def _xsend(wb, name):
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, download_name=name, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── Helpers PDF ───────────────────────────────────────────────────────────────

def _psetup(landscape=False):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.pagesizes import landscape as RL_L
    from reportlab.platypus import SimpleDocTemplate
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=RL_L(A4) if landscape else A4,
                             leftMargin=36, rightMargin=36,
                             topMargin=36, bottomMargin=36)
    return doc, buf


def _ptitle(title, subtitle=''):
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
    r, g, b = _PRIM
    pc = colors.Color(r/255, g/255, b/255)
    ts = ParagraphStyle('T', fontSize=14, fontName='Helvetica-Bold',
                         textColor=colors.white, alignment=1)
    ss = ParagraphStyle('S', fontSize=9,  fontName='Helvetica',
                         textColor=colors.white, alignment=1)
    ds = ParagraphStyle('D', fontSize=7,  fontName='Helvetica',
                         textColor=colors.white, alignment=2)
    rows = [[Paragraph(title, ts)]]
    if subtitle: rows.append([Paragraph(subtitle, ss)])
    rows.append([Paragraph(
        f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}', ds)])
    t = Table(rows, colWidths=['100%'])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), pc),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 14),
        ('RIGHTPADDING', (0, 0), (-1, -1), 14),
    ]))
    return [t, Spacer(1, 0.35*cm)]


def _ptbl(data, widths, zebra=True):
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle
    r, g, b = _PRIM
    pc = colors.Color(r/255, g/255, b/255)
    zc = colors.Color(250/255, 228/255, 235/255)
    t  = Table(data, colWidths=widths, repeatRows=1)
    st = [
        ('BACKGROUND',    (0, 0), (-1, 0),  pc),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0),  8),
        ('ALIGN',         (0, 0), (-1, 0),  'CENTER'),
        ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',      (0, 1), (-1, -1), 8),
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.Color(.75, .75, .75)),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING',   (0, 0), (-1, -1), 5),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 5),
    ]
    if zebra:
        for i in range(1, len(data)):
            if i % 2 == 0: st.append(('BACKGROUND', (0, i), (-1, i), zc))
    t.setStyle(TableStyle(st))
    return t


def _psend(doc, buf, story, name):
    doc.build(story); buf.seek(0)
    return send_file(buf, download_name=name, as_attachment=True,
                     mimetype='application/pdf')


# ═══════════════════════════════════════════════════════════════════════════════
# Índice
# ═══════════════════════════════════════════════════════════════════════════════

@reportes_bp.route('/')
@login_required
def index():
    _guard()
    return render_template('reportes/index.html',
                           cursos=_cursos(), alumnos=_alumnos())


# ═══════════════════════════════════════════════════════════════════════════════
# R1 — Alumnos por Curso
# ═══════════════════════════════════════════════════════════════════════════════

@reportes_bp.route('/r1')
@login_required
def r1():
    _guard()
    cur_id  = request.args.get('cur_id',  type=int)
    orden   = request.args.get('orden',   'nombre')
    formato = request.args.get('formato', 'html')
    curso   = Curso.query.get(cur_id) if cur_id else None
    filas   = []

    if cur_id:
        q = _ins_q(cur_id)
        if   orden == 'ci':   q = q.order_by(Alumno.ci)
        elif orden == 'sexo': q = q.order_by(Alumno.masculino.desc(),
                                              Alumno.paterno, Alumno.nombre)
        else:                 q = q.order_by(Alumno.paterno, Alumno.nombre)
        for i, ins in enumerate(q.all(), 1):
            a = ins.alumno
            filas.append({
                'n': i, 'nombre': a.nombre_completo, 'ci': a.ci or '—',
                'nac': a.nacimiento.strftime('%d/%m/%Y') if a.nacimiento else '—',
                'sexo': 'M' if a.masculino else 'F', 'email': a.email or '—',
                'estado': 'Inscrito' if ins.inscrito else 'Reserva',
                'desc': f'{ins.descuento}%' if ins.descuento else '—',
            })

    if formato == 'excel': return _r1_xl(filas, curso)
    if formato == 'pdf':   return _r1_pdf(filas, curso)
    return render_template('reportes/r1.html',
                           filas=filas, cursos=_cursos(),
                           cur_id=cur_id, curso=curso, orden=orden)


def _r1_xl(filas, curso):
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = 'Alumnos por Curso'
    s = _xs()
    _xh(ws, ['#', 'Apellidos y Nombre', 'CI', 'Nacimiento',
              'Sexo', 'Email', 'Estado', 'Descuento'], s)
    for f in filas:
        ws.append([f['n'], f['nombre'], f['ci'], f['nac'],
                   f['sexo'], f['email'], f['estado'], f['desc']])
        _xrow(ws, s)
    _xz(ws, s); _xw(ws)
    return _xsend(wb, f'r1_{(_nc(curso) or "todos").replace(" ","_")}.xlsx')


def _r1_pdf(filas, curso):
    from reportlab.lib.units import cm
    doc, buf = _psetup()
    story = _ptitle('Alumnos por Curso', _nc(curso))
    enc   = [['#', 'Nombre', 'CI', 'Nac.', 'Sexo', 'Email', 'Estado', 'Desc.']]
    data  = enc + [[f['n'], f['nombre'], f['ci'], f['nac'],
                    f['sexo'], f['email'], f['estado'], f['desc']] for f in filas]
    story.append(_ptbl(data, [.5*cm, 4.5*cm, 1.5*cm, 2*cm,
                               1*cm, 3.5*cm, 1.5*cm, 1*cm]))
    return _psend(doc, buf, story,
                  f'r1_{(_nc(curso) or "todos").replace(" ","_")}.pdf')


# ═══════════════════════════════════════════════════════════════════════════════
# R2 — Alumnos con Calificaciones
# ═══════════════════════════════════════════════════════════════════════════════

@reportes_bp.route('/r2')
@login_required
def r2():
    _guard()
    cur_id  = request.args.get('cur_id',  type=int)
    mat_id  = request.args.get('mat_id',  type=int)
    orden   = request.args.get('orden',   'nombre')
    formato = request.args.get('formato', 'html')
    curso   = Curso.query.get(cur_id) if cur_id else None
    materias = []; filas = []

    if cur_id:
        mat_ids  = [a.mat_id for a in Asignado.query.filter_by(cur_id=cur_id)]
        materias = (Materia.query.filter(Materia.id.in_(mat_ids))
                    .order_by(Materia.materia).all())
        q = (db.session.query(Nota, Alumno, Materia)
             .join(Inscrito, Nota.ins_id == Inscrito.id)
             .join(Alumno,   Inscrito.alu_id == Alumno.id)
             .join(Materia,  Nota.mat_id == Materia.id)
             .filter(Inscrito.cur_id == cur_id))
        if mat_id: q = q.filter(Nota.mat_id == mat_id)
        if   orden == 'nota':   q = q.order_by(Nota.nota_final.desc(), Alumno.paterno)
        elif orden == 'estado': q = q.order_by(Nota.aprobado.desc(), Nota.nota_final.desc())
        else:                   q = q.order_by(Alumno.paterno, Alumno.nombre, Materia.materia)
        for i, (n, a, m) in enumerate(q.all(), 1):
            filas.append({
                'n': i, 'alumno': a.nombre_completo, 'mat': m.materia,
                'n1': n.nota1, 'n2': n.nota2, 'n3': n.nota3,
                'nf': float(n.nota_final),
                'apr': 'Aprobado' if n.aprobado else 'Reprobado',
                'obs': n.obs or '',
            })

    if formato == 'excel': return _r2_xl(filas, curso)
    if formato == 'pdf':   return _r2_pdf(filas, curso)
    return render_template('reportes/r2.html',
                           filas=filas, cursos=_cursos(), materias=materias,
                           cur_id=cur_id, mat_id=mat_id, curso=curso, orden=orden)


def _r2_xl(filas, curso):
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = 'Calificaciones'
    s = _xs()
    _xh(ws, ['#', 'Alumno', 'Materia', 'P1', 'P2', 'P3',
              'Promedio', 'Estado', 'Observaciones'], s)
    for f in filas:
        ws.append([f['n'], f['alumno'], f['mat'], f['n1'], f['n2'], f['n3'],
                   f['nf'], f['apr'], f['obs']])
        _xrow(ws, s, fill=s['gf'] if f['apr'] == 'Aprobado' else s['rf'])
    _xw(ws)
    return _xsend(wb, f'r2_{(_nc(curso) or "todos").replace(" ","_")}.xlsx')


def _r2_pdf(filas, curso):
    from reportlab.lib.units import cm
    doc, buf = _psetup(landscape=True)
    story = _ptitle('Calificaciones por Curso', _nc(curso))
    enc   = [['#', 'Alumno', 'Materia', 'P1', 'P2', 'P3', 'Prom.', 'Estado', 'Obs.']]
    data  = enc + [[f['n'], f['alumno'], f['mat'], f['n1'], f['n2'], f['n3'],
                    f['nf'], f['apr'], f['obs']] for f in filas]
    story.append(_ptbl(data, [.5*cm, 4.5*cm, 3.5*cm, 1*cm,
                               1*cm, 1*cm, 1.2*cm, 2*cm, 3*cm]))
    return _psend(doc, buf, story,
                  f'r2_{(_nc(curso) or "todos").replace(" ","_")}.pdf')


# ═══════════════════════════════════════════════════════════════════════════════
# R3 — Estado de Pagos por Curso
# ═══════════════════════════════════════════════════════════════════════════════

@reportes_bp.route('/r3')
@login_required
def r3():
    _guard()
    cur_id  = request.args.get('cur_id',  type=int)
    orden   = request.args.get('orden',   'nombre')
    formato = request.args.get('formato', 'html')
    curso   = Curso.query.get(cur_id) if cur_id else None
    filas   = []

    if cur_id:
        rows = []
        for ins in (_ins_q(cur_id)
                    .order_by(Alumno.paterno, Alumno.nombre).all()):
            ps   = ins.pagos.all()
            tot  = sum(p.cuota for p in ps)
            pag  = sum(p.cuota for p in ps if p.pagado)
            pend = tot - pag; cp = len(ps); cpg = sum(1 for p in ps if p.pagado)
            rows.append({
                'nombre': ins.alumno.nombre_completo, 'ci': ins.alumno.ci or '—',
                'ct': cp, 'cpg': cpg, 'tot': tot, 'pag': pag, 'pend': pend,
                'estado': ('Al día'    if pend == 0 and cp > 0
                            else 'Con deuda' if pend > 0 else 'Sin plan'),
            })
        if orden == 'deuda':    rows.sort(key=lambda x: x['pend'], reverse=True)
        elif orden == 'estado': rows.sort(key=lambda x: x['estado'])
        for i, r in enumerate(rows, 1): r['n'] = i; filas.append(r)

    if formato == 'excel': return _r3_xl(filas, curso)
    if formato == 'pdf':   return _r3_pdf(filas, curso)
    return render_template('reportes/r3.html',
                           filas=filas, cursos=_cursos(),
                           cur_id=cur_id, curso=curso, orden=orden)


def _r3_xl(filas, curso):
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = 'Estado de Pagos'
    s = _xs()
    _xh(ws, ['#', 'Alumno', 'CI', 'Cuotas Tot.', 'Cuotas Pag.',
              'Total Bs.', 'Pagado Bs.', 'Pendiente Bs.', 'Estado'], s)
    for f in filas:
        ws.append([f['n'], f['nombre'], f['ci'], f['ct'], f['cpg'],
                   round(f['tot'], 2), round(f['pag'], 2),
                   round(f['pend'], 2), f['estado']])
        _xrow(ws, s, fill=s['gf'] if f['estado'] == 'Al día' else s['rf'])
    _xw(ws)
    return _xsend(wb, f'r3_{(_nc(curso) or "todos").replace(" ","_")}.xlsx')


def _r3_pdf(filas, curso):
    from reportlab.lib.units import cm
    doc, buf = _psetup(landscape=True)
    story = _ptitle('Estado de Pagos por Curso', _nc(curso))
    enc   = [['#', 'Alumno', 'CI', 'C.Tot', 'C.Pag',
               'Total Bs.', 'Pagado Bs.', 'Pendiente Bs.', 'Estado']]
    data  = enc + [[f['n'], f['nombre'], f['ci'], f['ct'], f['cpg'],
                    f"{f['tot']:.2f}", f"{f['pag']:.2f}",
                    f"{f['pend']:.2f}", f['estado']] for f in filas]
    story.append(_ptbl(data, [.5*cm, 4.5*cm, 1.5*cm, 1.2*cm, 1.2*cm,
                               2*cm, 2*cm, 2.5*cm, 2*cm]))
    return _psend(doc, buf, story,
                  f'r3_{(_nc(curso) or "todos").replace(" ","_")}.pdf')


# ═══════════════════════════════════════════════════════════════════════════════
# R4 — Estado de Cuenta de un Alumno
# ═══════════════════════════════════════════════════════════════════════════════

@reportes_bp.route('/r4')
@login_required
def r4():
    _guard()
    alu_id  = request.args.get('alu_id',  type=int)
    formato = request.args.get('formato', 'html')
    alumno  = Alumno.query.get(alu_id) if alu_id else None
    ins     = Inscrito.query.filter_by(alu_id=alu_id).first() if alu_id else None
    pagos   = []; resumen = {}

    if ins:
        for p in Pago.query.filter_by(ins_id=ins.id).order_by(Pago.nro_cuota).all():
            pagos.append({
                'nro': p.nro_cuota, 'cuota': p.cuota, 'pagado': p.pagado,
                'metodo': p.metodo_pago or '—',
                'fecha':  p.fecha_pago.strftime('%d/%m/%Y') if p.fecha_pago else '—',
                'ref':    p.referencia_pago or '—', 'obs': p.obs or '',
            })
        tot = sum(p['cuota'] for p in pagos); pag = sum(p['cuota'] for p in pagos if p['pagado'])
        resumen = {'total': tot, 'pagado': pag, 'pendiente': tot - pag,
                   'cuotas': len(pagos),
                   'cpag': sum(1 for p in pagos if p['pagado'])}

    if formato == 'excel': return _r4_xl(alumno, ins, pagos, resumen)
    if formato == 'pdf':   return _r4_pdf(alumno, ins, pagos, resumen)
    return render_template('reportes/r4.html',
                           alumno=alumno, ins=ins, pagos=pagos,
                           resumen=resumen, alu_id=alu_id, alumnos=_alumnos())


def _r4_xl(alumno, ins, pagos, resumen):
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = 'Estado de Cuenta'
    s = _xs()
    ws.append(['Alumno:',    alumno.nombre_completo if alumno else ''])
    ws.append(['CI:',        str(alumno.ci) if alumno and alumno.ci else ''])
    ws.append(['Curso:',     _nc(ins.curso) if ins and ins.curso else ''])
    ws.append(['Descuento:', f'{ins.descuento}%' if ins and ins.descuento else '0%'])
    ws.append([])
    _xh(ws, ['Cuota #', 'Monto Bs.', 'Estado', 'Método',
              'Fecha Pago', 'Referencia', 'Observaciones'], s)
    for p in pagos:
        ws.append([p['nro'], round(p['cuota'], 2),
                   'Pagado' if p['pagado'] else 'Pendiente',
                   p['metodo'], p['fecha'], p['ref'], p['obs']])
        _xrow(ws, s, fill=s['gf'] if p['pagado'] else s['am'])
    ws.append([])
    ws.append(['', 'TOTAL:', resumen.get('total', 0), 'PAGADO:',
               resumen.get('pagado', 0), 'PENDIENTE:', resumen.get('pendiente', 0)])
    for c in ws[ws.max_row]: c.font = s['bld']
    _xw(ws)
    fn = alumno.nombre_completo.replace(' ', '_') if alumno else 'alumno'
    return _xsend(wb, f'r4_{fn}.xlsx')


def _r4_pdf(alumno, ins, pagos, resumen):
    from reportlab.lib.units import cm
    from reportlab.platypus import Spacer, Table, TableStyle
    doc, buf = _psetup()
    story = _ptitle('Estado de Cuenta del Alumno',
                     alumno.nombre_completo if alumno else '')
    info = [['Alumno:', alumno.nombre_completo if alumno else '—'],
            ['CI:',     str(alumno.ci) if alumno and alumno.ci else '—'],
            ['Curso:',  _nc(ins.curso) if ins and ins.curso else '—'],
            ['Descuento:', f'{ins.descuento}%' if ins and ins.descuento else '0%']]
    it = Table(info, colWidths=[3*cm, 12*cm])
    it.setStyle(TableStyle([('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                              ('FONTSIZE', (0, 0), (-1, -1), 9),
                              ('TOPPADDING', (0, 0), (-1, -1), 3),
                              ('BOTTOMPADDING', (0, 0), (-1, -1), 3)]))
    story += [it, Spacer(1, .3*cm)]
    enc   = [['Cuota #', 'Monto Bs.', 'Estado', 'Método', 'Fecha', 'Referencia']]
    data  = enc + [[p['nro'], f"{p['cuota']:.2f}",
                    'Pagado' if p['pagado'] else 'Pendiente',
                    p['metodo'], p['fecha'], p['ref']] for p in pagos]
    story.append(_ptbl(data, [1.5*cm, 2.5*cm, 2*cm, 2.5*cm, 2.5*cm, 4*cm]))
    story.append(Spacer(1, .3*cm))
    res  = [['Total Bs.', 'Pagado Bs.', 'Pendiente Bs.', 'Cuotas', 'C. Pagadas'],
            [f"{resumen.get('total', 0):.2f}", f"{resumen.get('pagado', 0):.2f}",
             f"{resumen.get('pendiente', 0):.2f}",
             resumen.get('cuotas', 0), resumen.get('cpag', 0)]]
    story.append(_ptbl(res, [3*cm, 3*cm, 3.5*cm, 2*cm, 3.5*cm], zebra=False))
    fn = alumno.nombre_completo.replace(' ', '_') if alumno else 'alumno'
    return _psend(doc, buf, story, f'r4_{fn}.pdf')


# ═══════════════════════════════════════════════════════════════════════════════
# R5 — Top 3 mejores promedios
# ═══════════════════════════════════════════════════════════════════════════════

@reportes_bp.route('/r5')
@login_required
def r5():
    _guard()
    cur_id  = request.args.get('cur_id',  type=int)
    formato = request.args.get('formato', 'html')
    cursos_sel = [Curso.query.get(cur_id)] if cur_id else _cursos()
    filas = []

    for curso in cursos_sel:
        if not curso: continue
        subq = (db.session.query(
                    Inscrito.id.label('iid'),
                    func.avg(Nota.nota_final).label('prom'),
                    func.sum(db.case((Nota.aprobado == True, 1), else_=0)).label('apr'),
                    func.sum(db.case((Nota.aprobado == False, 1), else_=0)).label('rep'),
                )
                .join(Nota, Nota.ins_id == Inscrito.id)
                .filter(Inscrito.cur_id == curso.id)
                .group_by(Inscrito.id)
                .order_by(func.avg(Nota.nota_final).desc())
                .limit(3).all())
        for pos, row in enumerate(subq, 1):
            ins = Inscrito.query.get(row.iid)
            if ins and ins.alumno:
                filas.append({
                    'curso': _nc(curso), 'pos': pos,
                    'alumno': ins.alumno.nombre_completo,
                    'prom': round(float(row.prom), 2),
                    'apr': row.apr, 'rep': row.rep,
                })

    if formato == 'excel': return _r5_xl(filas)
    if formato == 'pdf':   return _r5_pdf(filas)
    return render_template('reportes/r5.html',
                           filas=filas, cursos=_cursos(), cur_id=cur_id)


def _r5_xl(filas):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    wb = Workbook(); ws = wb.active; ws.title = 'Top 3'
    s = _xs(); medals = {1: 'FFFFD700', 2: 'FFC0C0C0', 3: 'FFCD7F32'}
    _xh(ws, ['Curso', 'Puesto', 'Alumno', 'Promedio',
              'Materias Aprobadas', 'Materias Reprobadas'], s)
    for f in filas:
        ws.append([f['curso'], f'#{f["pos"]}', f['alumno'],
                   f['prom'], f['apr'], f['rep']])
        fill = PatternFill('solid', fgColor=medals.get(f['pos'], 'FFFFFFFF'))
        for c in ws[ws.max_row]:
            c.fill = fill; c.font = s['bf']
            c.border = s['brd']; c.alignment = s['lft']
    _xw(ws)
    return _xsend(wb, 'r5_top3.xlsx')


def _r5_pdf(filas):
    from reportlab.lib.units import cm
    doc, buf = _psetup()
    story = _ptitle('Top 3 Mejores Promedios por Curso')
    enc   = [['Curso', 'Puesto', 'Alumno', 'Promedio', 'Aprob.', 'Reprobadas']]
    data  = enc + [[f['curso'], f'#{f["pos"]}', f['alumno'],
                    f['prom'], f['apr'], f['rep']] for f in filas]
    story.append(_ptbl(data, [5*cm, 1.2*cm, 4.5*cm, 1.8*cm, 1.5*cm, 1.5*cm]))
    return _psend(doc, buf, story, 'r5_top3.pdf')


# ═══════════════════════════════════════════════════════════════════════════════
# R6 — Resumen Financiero por Curso  ★ adicional
# ═══════════════════════════════════════════════════════════════════════════════

@reportes_bp.route('/r6')
@login_required
def r6():
    _guard()
    formato = request.args.get('formato', 'html')
    filas   = []
    for c in _cursos():
        il = (Inscrito.query.filter(Inscrito.cur_id == c.id,
                db.or_(Inscrito.inscrito == True,
                       Inscrito.reserva  == True)).all())
        esp  = sum(sum(p.cuota for p in i.pagos.all()) for i in il)
        rec  = sum(sum(p.cuota for p in i.pagos.all() if p.pagado) for i in il)
        pend = esp - rec
        pct  = round(rec / esp * 100, 1) if esp else 0
        filas.append({
            'curso': _nc(c), 'alumnos': len(il),
            'esp': round(esp, 2), 'rec': round(rec, 2),
            'pend': round(pend, 2), 'pct': pct,
        })
    if formato == 'excel': return _r6_xl(filas)
    if formato == 'pdf':   return _r6_pdf(filas)
    return render_template('reportes/r6.html', filas=filas)


def _r6_xl(filas):
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = 'Resumen Financiero'
    s = _xs()
    _xh(ws, ['Curso', 'Alumnos', 'Total Esp. Bs.',
              'Recaudado Bs.', 'Pendiente Bs.', '% Cobrado'], s)
    for f in filas:
        ws.append([f['curso'], f['alumnos'],
                   f['esp'], f['rec'], f['pend'], f['pct']])
        fill = s['gf'] if f['pct'] >= 80 else (s['am'] if f['pct'] >= 50 else s['rf'])
        _xrow(ws, s, fill=fill)
    _xw(ws)
    return _xsend(wb, 'r6_resumen_financiero.xlsx')


def _r6_pdf(filas):
    from reportlab.lib.units import cm
    doc, buf = _psetup()
    story = _ptitle('Resumen Financiero por Curso')
    enc   = [['Curso', 'Alumnos', 'Esperado Bs.',
               'Recaudado Bs.', 'Pendiente Bs.', '% Cobrado']]
    data  = enc + [[f['curso'], f['alumnos'], f"{f['esp']:.2f}",
                    f"{f['rec']:.2f}", f"{f['pend']:.2f}",
                    f"{f['pct']}%"] for f in filas]
    story.append(_ptbl(data, [4.5*cm, 1.8*cm, 3*cm, 3*cm, 3*cm, 2*cm]))
    return _psend(doc, buf, story, 'r6_resumen_financiero.pdf')


# ═══════════════════════════════════════════════════════════════════════════════
# R7 — Promedios y Tasa de Aprobación por Materia  ★ adicional
# ═══════════════════════════════════════════════════════════════════════════════

@reportes_bp.route('/r7')
@login_required
def r7():
    _guard()
    cur_id  = request.args.get('cur_id',  type=int)
    orden   = request.args.get('orden',   'materia')
    formato = request.args.get('formato', 'html')
    curso   = Curso.query.get(cur_id) if cur_id else None
    filas   = []

    q = (db.session.query(
             Materia.materia,
             func.count(Nota.id).label('tot'),
             func.sum(db.case((Nota.aprobado == True, 1), else_=0)).label('apr'),
             func.avg(Nota.nota_final).label('prom'),
             func.min(Nota.nota_final).label('mn'),
             func.max(Nota.nota_final).label('mx'),
         )
         .join(Nota,     Nota.mat_id == Materia.id)
         .join(Inscrito, Nota.ins_id == Inscrito.id))
    if cur_id: q = q.filter(Inscrito.cur_id == cur_id)
    q = q.group_by(Materia.materia)
    if orden == 'promedio':
        q = q.order_by(func.avg(Nota.nota_final).desc())
    elif orden == 'tasa':
        q = q.order_by(
            (func.sum(db.case((Nota.aprobado == True, 1), else_=0)) /
             func.count(Nota.id)).desc())
    else:
        q = q.order_by(Materia.materia)

    for i, row in enumerate(q.all(), 1):
        rep = row.tot - row.apr
        pct = round(float(row.apr) / row.tot * 100, 1) if row.tot else 0
        filas.append({
            'n': i, 'mat': row.materia, 'tot': row.tot,
            'apr': row.apr, 'rep': rep, 'pct': pct,
            'prom': round(float(row.prom), 2),
            'mn': float(row.mn), 'mx': float(row.mx),
        })

    if formato == 'excel': return _r7_xl(filas, curso)
    if formato == 'pdf':   return _r7_pdf(filas, curso)
    return render_template('reportes/r7.html',
                           filas=filas, cursos=_cursos(),
                           cur_id=cur_id, curso=curso, orden=orden)


def _r7_xl(filas, curso):
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = 'Tasa Aprobación'
    s = _xs()
    _xh(ws, ['#', 'Materia', 'Total', 'Aprob.', 'Reprobados',
              '% Aprob.', 'Promedio', 'Mín.', 'Máx.'], s)
    for f in filas:
        ws.append([f['n'], f['mat'], f['tot'], f['apr'], f['rep'],
                   f['pct'], f['prom'], f['mn'], f['mx']])
        fill = s['gf'] if f['pct'] >= 70 else (s['am'] if f['pct'] >= 50 else s['rf'])
        _xrow(ws, s, fill=fill)
    _xw(ws)
    return _xsend(wb, f'r7_{(_nc(curso) or "todos").replace(" ","_")}.xlsx')


def _r7_pdf(filas, curso):
    from reportlab.lib.units import cm
    doc, buf = _psetup(landscape=True)
    story = _ptitle('Promedios y Tasa de Aprobación por Materia',
                     _nc(curso) or 'Todos los cursos')
    enc   = [['#', 'Materia', 'Total', 'Aprob.', 'Reprobados',
               '% Aprob.', 'Prom.', 'Mín.', 'Máx.']]
    data  = enc + [[f['n'], f['mat'], f['tot'], f['apr'], f['rep'],
                    f"{f['pct']}%", f['prom'], f['mn'], f['mx']] for f in filas]
    story.append(_ptbl(data, [.6*cm, 4*cm, 1.2*cm, 1.3*cm, 1.5*cm,
                               1.5*cm, 2*cm, 1.2*cm, 1.2*cm]))
    return _psend(doc, buf, story,
                  f'r7_{(_nc(curso) or "todos").replace(" ","_")}.pdf')


# ═══════════════════════════════════════════════════════════════════════════════
# R8 — Alumnos con Deuda o Abandono  ★ adicional
# ═══════════════════════════════════════════════════════════════════════════════

@reportes_bp.route('/r8')
@login_required
def r8():
    _guard()
    cur_id  = request.args.get('cur_id',  type=int)
    tipo    = request.args.get('tipo',    'deuda')
    formato = request.args.get('formato', 'html')
    curso   = Curso.query.get(cur_id) if cur_id else None
    filas   = []

    q = Inscrito.query.join(Alumno, Inscrito.alu_id == Alumno.id)
    if cur_id: q = q.filter(Inscrito.cur_id == cur_id)

    for ins in q.order_by(Alumno.paterno, Alumno.nombre).all():
        ps   = ins.pagos.all()
        pend = sum(p.cuota for p in ps if not p.pagado)
        ed   = pend > 0; ea = ins.abandono
        if tipo == 'deuda'    and not ed:         continue
        if tipo == 'abandono' and not ea:         continue
        if tipo == 'ambos'    and not (ed or ea): continue
        filas.append({
            'alumno': ins.alumno.nombre_completo, 'ci': ins.alumno.ci or '—',
            'curso':  _nc(ins.curso) if ins.curso else '—',
            'deuda':  round(pend, 2),
            'abandono': 'Sí' if ea else 'No',
            'obs': ins.obs or '',
        })

    if formato == 'excel': return _r8_xl(filas, tipo)
    if formato == 'pdf':   return _r8_pdf(filas, tipo)
    return render_template('reportes/r8.html',
                           filas=filas, cursos=_cursos(),
                           cur_id=cur_id, tipo=tipo, curso=curso)


def _r8_xl(filas, tipo):
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = 'Deuda/Abandono'
    s = _xs()
    _xh(ws, ['Alumno', 'CI', 'Curso', 'Deuda Bs.', 'Abandono', 'Obs.'], s)
    for f in filas:
        ws.append([f['alumno'], f['ci'], f['curso'],
                   f['deuda'], f['abandono'], f['obs']])
        _xrow(ws, s, fill=s['rf'] if f['abandono'] == 'Sí' else s['am'])
    _xw(ws)
    return _xsend(wb, f'r8_{tipo}.xlsx')


def _r8_pdf(filas, tipo):
    from reportlab.lib.units import cm
    doc, buf = _psetup()
    story = _ptitle('Alumnos con Deuda / Abandono')
    enc   = [['Alumno', 'CI', 'Curso', 'Deuda Bs.', 'Abandono', 'Obs.']]
    data  = enc + [[f['alumno'], f['ci'], f['curso'],
                    f"{f['deuda']:.2f}", f['abandono'], f['obs']] for f in filas]
    story.append(_ptbl(data, [4.5*cm, 1.5*cm, 3.5*cm, 2*cm, 1.5*cm, 3*cm]))
    return _psend(doc, buf, story, f'r8_{tipo}.pdf')


# ═══════════════════════════════════════════════════════════════════════════════
# AJAX — materias de un curso (para filtro dinámico en R2)
# ═══════════════════════════════════════════════════════════════════════════════

@reportes_bp.route('/materias')
@login_required
def materias_ajax():
    from flask import jsonify
    _guard()
    cur_id = request.args.get('cur_id', type=int)
    if not cur_id:
        return jsonify([])
    mat_ids  = [a.mat_id for a in Asignado.query.filter_by(cur_id=cur_id)]
    materias = (Materia.query.filter(Materia.id.in_(mat_ids))
                .order_by(Materia.materia).all())
    return jsonify([{'id': m.id, 'materia': m.materia} for m in materias])
